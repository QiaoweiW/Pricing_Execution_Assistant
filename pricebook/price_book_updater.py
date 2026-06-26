"""
Update Price Book workbooks from an Oracle price-adjustment extract.

Given one Oracle extract (CSV, ``priceadjs`` schema) and one-or-more customer
"Price Book" workbooks (.xlsx), fill each workbook's **Old Price** / **New
Price** columns from the extract's ``adjustmentamount`` for two operator-chosen
snapshots (``adjustmentstartdate``):

    Old Price = adjustmentamount where adjustmentstartdate == Old period
    New Price = adjustmentamount where adjustmentstartdate == New period

A Price Book row is matched to the extract on three normalized keys:

    Darigold Item Number  <->  itemname
    Customer Site Name     <->  shiptositename   (from the workbook's header block)
    Pricing UOM            <->  pricinguom

Only matched rows are touched; ``Price Change`` is recomputed (New - Old) for
updated rows and the updated rows are highlighted light-yellow. Everything else
in the workbook (logo, merged header, number formats, untouched rows) is left
byte-for-byte intact because we edit in place with openpyxl rather than
rebuilding the sheet.

Pure, dependency-light, and Streamlit-free so it is unit-testable: the only
third-party imports are pandas and (lazily) openpyxl, both already in
requirements.txt.
"""
from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd

# ── Oracle extract ───────────────────────────────────────────────────────────
# Columns the extract MUST carry (lower-cased on load). Mirrors priceadjs.
_EXTRACT_REQUIRED = (
    "itemname", "shiptositename", "pricinguom",
    "adjustmentamount", "adjustmentstartdate",
)

# ── Price Book workbook layout (matched by header text, never by fixed cell) ──
_PB_SITE_LABEL = "customer site name"   # header-block label; value is to its right
_PB_ITEM       = "darigold item number"
_PB_UOM        = "pricing uom"
_PB_OLD        = "old price"
_PB_NEW        = "new price"
_PB_CHANGE     = "price change"

# Light-yellow fill (ARGB) applied to every row we update.
_HIGHLIGHT_ARGB = "FFFFFF99"


class PriceBookUpdateError(RuntimeError):
    """Raised when an input file can't be parsed / is missing required columns."""


@dataclass
class UpdateReport:
    """Per-workbook outcome — the UI renders this and offers the bytes."""
    file_name: str
    ok: bool
    level: str                                  # success | warning | error
    message: str
    site_name: Optional[str] = None
    site_found: bool = False
    rows_total: int = 0
    rows_updated: int = 0
    old_set: int = 0
    new_set: int = 0
    unmatched: list[str] = field(default_factory=list)   # "item (uom)" not found
    workbook_bytes: Optional[bytes] = None      # updated (or original) .xlsx


# ── Normalization helpers ────────────────────────────────────────────────────

def _norm_text(v: object) -> str:
    """Collapse NBSP/whitespace and strip — for labels, headers, site names."""
    return re.sub(r"\s+", " ", str(v).replace("\xa0", " ")).strip()


def _norm_key(v: object) -> str:
    """Case-insensitive match key (site / UOM)."""
    return _norm_text(v).upper()


def _norm_item(v: object) -> str:
    """Item number as a clean digit string across int / float / str inputs.

    Excel stores item numbers as ints (``340021``); a stray float (``340021.0``)
    or string with whitespace must collapse to the same key as the extract's
    string ``"340021"``.
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return re.sub(r"\.0$", "", _norm_text(v))


# ── Extract → lookup ─────────────────────────────────────────────────────────

def load_oracle_extract(file_bytes: bytes) -> pd.DataFrame:
    """Parse the Oracle extract CSV (all-strings) and validate its columns."""
    try:
        df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, keep_default_na=False)
    except Exception as exc:  # noqa: BLE001
        raise PriceBookUpdateError(f"Could not read the Oracle extract CSV: {exc}") from exc
    df.columns = [_norm_text(c).lower() for c in df.columns]
    missing = [c for c in _EXTRACT_REQUIRED if c not in df.columns]
    if missing:
        raise PriceBookUpdateError(
            "Oracle extract is missing required column(s): " + ", ".join(missing)
        )
    return df


def available_periods(extract_df: pd.DataFrame) -> list[pd.Timestamp]:
    """Sorted distinct ``adjustmentstartdate`` snapshots (tz-aware, UTC)."""
    ts = pd.to_datetime(extract_df["adjustmentstartdate"], utc=True, errors="coerce")
    return sorted(ts.dropna().unique())


def format_period(ts: pd.Timestamp) -> str:
    """Human label for a period dropdown / caption."""
    return pd.Timestamp(ts).strftime("%Y-%m-%d %H:%M:%S %Z")


def build_lookup(extract_df: pd.DataFrame) -> tuple[dict, set]:
    """Build ``{(item, site, uom, period_ts): amount}`` + the set of known sites.

    All rows are considered **regardless of ``status``** — reference/draft
    extracts can be entirely ``U`` (Update-Pending) rather than ``S``, and the
    user expects those to match. Amounts are unique at the (item, site, uom,
    startdate) grain in practice; if a future extract ever carries the same key
    twice, the latest ``last_update_date`` / ``batchno`` wins (deterministic).
    """
    df = extract_df
    work = pd.DataFrame({
        "item": df["itemname"].map(_norm_item),
        "site": df["shiptositename"].map(_norm_key),
        "uom":  df["pricinguom"].map(_norm_key),
        "ts":   pd.to_datetime(df["adjustmentstartdate"], utc=True, errors="coerce"),
        "amt":  pd.to_numeric(df["adjustmentamount"], errors="coerce"),
    })
    # Tie-breakers for the (defensive) conflict case.
    work["_lud"] = (pd.to_datetime(df["last_update_date"], utc=True, errors="coerce")
                    if "last_update_date" in df.columns else pd.NaT)
    work["_batch"] = df["batchno"].astype(str) if "batchno" in df.columns else ""
    work = work.dropna(subset=["ts", "amt"])
    # Latest update / batch first, so the first row per key is the winner.
    work = work.sort_values(["_lud", "_batch"], ascending=[False, False], kind="stable")

    lookup: dict[tuple, float] = {}
    for item, site, uom, ts, amt in zip(
        work["item"], work["site"], work["uom"], work["ts"], work["amt"]
    ):
        lookup.setdefault((item, site, uom, ts), float(amt))
    sites = set(work["site"].unique())
    return lookup, sites


# ── Workbook parsing helpers ─────────────────────────────────────────────────

def _find_site_name(ws) -> Optional[str]:
    """Return the workbook's Customer Site Name from its header block.

    Scans the top rows for the label cell, then takes the first non-empty cell
    to its right — robust to the exact row/column the report uses.
    """
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15)):
        for idx, cell in enumerate(row):
            if cell.value is not None and _norm_text(cell.value).lower() == _PB_SITE_LABEL:
                for nxt in row[idx + 1:]:
                    if nxt.value not in (None, ""):
                        return _norm_text(nxt.value)
    return None


def _find_header(ws) -> Optional[tuple[int, dict[str, int]]]:
    """Locate the table header row; return ``(row_idx, {header_lower: col_idx})``."""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40)):
        headers = {
            _norm_text(c.value).lower(): c.column
            for c in row if c.value not in (None, "")
        }
        if _PB_ITEM in headers:
            return row[0].row, headers
    return None


# ── Shared update engine (used by the .xlsx and .pdf paths alike) ────────────

def _amount_for_periods(
    item: str, site: str, uom: str,
    periods: list, lookup: dict,
) -> tuple[Optional[float], int]:
    """Resolve a row's amount across the selected ``periods``.

    Returns ``(amount, n_matches)``. ``periods`` empty → ``(None, 0)`` ("leave
    as-is"). When more than one selected period has a row for this key, the
    **latest** period's amount wins (and ``n_matches`` reports the overlap so the
    caller can flag it).
    """
    if not periods:
        return None, 0
    matches = [(p, lookup[(item, site, uom, p)])
               for p in periods if (item, site, uom, p) in lookup]
    if not matches:
        return None, 0
    matches.sort(key=lambda pa: pa[0])          # by period, ascending
    return matches[-1][1], len(matches)         # latest period's amount wins


def _apply_updates_to_worksheet(
    ws, header_row: int, cols: dict[str, int], site: str,
    lookup: dict, old_periods: list, new_periods: list,
) -> tuple[int, int, int, int, list[str], int]:
    """Fill Old/New (+ recompute Change) for matched rows; highlight them.

    The single source of truth for the match → write → highlight behaviour. Both
    the in-place ``.xlsx`` path and the ``.pdf`` → cloned-template path call this
    once they have a populated worksheet, so the two never drift apart.

    ``old_periods`` / ``new_periods`` are **lists** of selected snapshots. An
    empty list means "leave that column as-is" (never looked up or written). A
    row counts as updated/highlighted iff at least one side was filled. When a
    row matches more than one selected period, the latest period wins.

    Returns ``(rows_total, rows_updated, old_set, new_set, unmatched, multi)``
    where ``multi`` is the count of rows that matched >1 selected period.
    """
    from openpyxl.styles import PatternFill

    item_c, uom_c = cols[_PB_ITEM], cols[_PB_UOM]
    old_c, new_c = cols[_PB_OLD], cols[_PB_NEW]
    change_c = cols.get(_PB_CHANGE)
    fill_lo, fill_hi = min(cols.values()), max(cols.values())
    highlight = PatternFill(start_color=_HIGHLIGHT_ARGB,
                            end_color=_HIGHLIGHT_ARGB, fill_type="solid")

    rows_total = rows_updated = old_set = new_set = multi = 0
    unmatched: list[str] = []
    for r in range(header_row + 1, ws.max_row + 1):
        item_raw = ws.cell(r, item_c).value
        if item_raw in (None, ""):
            continue  # spacer / blank row
        if _norm_text(item_raw).lower() == _PB_ITEM:
            continue  # a repeated header (e.g. between UOM blocks)
        rows_total += 1

        item = _norm_item(item_raw)
        uom = _norm_key(ws.cell(r, uom_c).value)
        old_amt, n_old = _amount_for_periods(item, site, uom, old_periods, lookup)
        new_amt, n_new = _amount_for_periods(item, site, uom, new_periods, lookup)
        if n_old > 1 or n_new > 1:
            multi += 1

        if old_amt is None and new_amt is None:
            unmatched.append(f"{item} ({uom})")
            continue
        if old_amt is not None:
            ws.cell(r, old_c).value = old_amt
            old_set += 1
        if new_amt is not None:
            ws.cell(r, new_c).value = new_amt
            new_set += 1
        rows_updated += 1

        # Price Change = New - Old, recomputed when both cells are numeric
        # post-update (so a one-sided match never writes a bogus delta).
        if change_c is not None:
            ocv, ncv = ws.cell(r, old_c).value, ws.cell(r, new_c).value
            if isinstance(ocv, (int, float)) and isinstance(ncv, (int, float)):
                ws.cell(r, change_c).value = ncv - ocv

        for c in range(fill_lo, fill_hi + 1):
            ws.cell(r, c).fill = highlight

    return rows_total, rows_updated, old_set, new_set, unmatched, multi


def _build_update_report(
    file_name: str, site_raw: str, stats: tuple, workbook_bytes: bytes,
) -> UpdateReport:
    """Compose the success ``UpdateReport`` from the engine's stats tuple."""
    rows_total, rows_updated, old_set, new_set, unmatched, multi = stats
    msg = (f"Updated {rows_updated} of {rows_total} rows "
           f"(Old set: {old_set}, New set: {new_set}).")
    if unmatched:
        msg += f" {len(unmatched)} row(s) had no match in the extract."
    if multi:
        msg += f" {multi} row(s) matched multiple selected periods (latest used)."
    return UpdateReport(
        file_name, True, "success", msg,
        site_name=site_raw, site_found=True,
        rows_total=rows_total, rows_updated=rows_updated,
        old_set=old_set, new_set=new_set, unmatched=unmatched,
        workbook_bytes=workbook_bytes,
    )


# ── Per-workbook update ──────────────────────────────────────────────────────

def update_price_book(
    file_name: str,
    workbook_bytes: bytes,
    lookup: dict,
    known_sites: set,
    old_periods: list,
    new_periods: list,
) -> UpdateReport:
    """Fill Old/New Price for matched rows in one workbook; return the report.

    Edits in place so all formatting survives. When the workbook's site isn't in
    the extract, nothing is changed and a warning report is returned (the file
    still rides along in the output, unmodified).
    """
    from openpyxl import load_workbook  # lazy: keep the page import-safe

    try:
        wb = load_workbook(io.BytesIO(workbook_bytes))
    except Exception as exc:  # noqa: BLE001
        return UpdateReport(file_name, False, "error",
                            f"Could not open workbook: {exc}")
    ws = wb.active

    site_raw = _find_site_name(ws)
    if not site_raw:
        return UpdateReport(file_name, False, "error",
                            "Couldn't find a 'Customer Site Name' in the header block.",
                            workbook_bytes=workbook_bytes)
    site = _norm_key(site_raw)
    if site not in known_sites:
        return UpdateReport(
            file_name, True, "warning",
            f"Site '{site_raw}' was not found in the Oracle extract — no rows updated.",
            site_name=site_raw, site_found=False, workbook_bytes=workbook_bytes,
        )

    header = _find_header(ws)
    if header is None:
        return UpdateReport(file_name, False, "error",
                            "Couldn't find the 'Darigold Item Number' header row.",
                            site_name=site_raw, site_found=True,
                            workbook_bytes=workbook_bytes)
    header_row, cols = header
    missing = [c for c in (_PB_ITEM, _PB_UOM, _PB_OLD, _PB_NEW) if c not in cols]
    if missing:
        return UpdateReport(file_name, False, "error",
                            f"Workbook is missing column(s): {missing}",
                            site_name=site_raw, site_found=True,
                            workbook_bytes=workbook_bytes)

    stats = _apply_updates_to_worksheet(
        ws, header_row, cols, site, lookup, old_periods, new_periods)
    out = io.BytesIO()
    wb.save(out)
    return _build_update_report(file_name, site_raw, stats, out.getvalue())


# ── PDF price book → cloned Excel template ───────────────────────────────────
#
# PDFs have no editable cells, so (per the agreed design) we extract the table,
# clone the official Excel *template* for fidelity (logo, fonts, header block),
# write the rows in, and run the SAME update engine as the .xlsx path. The
# output is an .xlsx, not a PDF.


def _search_header(text: str, pattern: str) -> Optional[str]:
    m = re.search(pattern, text)
    return m.group(1).strip() if m else None


def _extract_pdf_price_book(pdf_bytes: bytes) -> tuple[dict, list[dict]]:
    """Extract ``(header_info, rows)`` from a digital Price Book PDF.

    ``header_info`` = ``{customer, site, site_no}`` (from page-1 text); ``rows``
    is a list of ``{canonical_header: value}`` dicts with wrapped text de-wrapped.
    Repeated per-page column headers are skipped; only the data rows survive.
    """
    import pdfplumber  # lazy: pdfplumber is already a project dependency

    header_info: dict = {"customer": None, "site": None, "site_no": None}
    header_cols: Optional[list[str]] = None
    rows: list[dict] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_no, page in enumerate(pdf.pages):
            if page_no == 0:
                txt = page.extract_text() or ""
                header_info["customer"] = _search_header(txt, r"Customer\s+(.*?)\s+Report Submission Date")
                header_info["site"] = _search_header(txt, r"Customer Site Name\s+(.*?)\s+Frequency")
                header_info["site_no"] = _search_header(txt, r"Customer Site No\s+(\S+)")
            for table in page.extract_tables() or []:
                for raw in table or []:
                    cells = [_norm_text(c) if c is not None else "" for c in raw]
                    if not cells:
                        continue
                    if cells[0].lower() == _PB_ITEM:          # (repeated) header row
                        header_cols = [c.lower() for c in cells]
                        continue
                    if header_cols is None or not cells[0].strip():
                        continue                              # pre-header / blank
                    rows.append({
                        header_cols[i]: cells[i]
                        for i in range(min(len(header_cols), len(cells)))
                    })
    return header_info, rows


def _coerce_pdf_value(canon_col: str, raw: str):
    """Type a PDF cell to match the Excel template's expected column type."""
    raw = (raw or "").strip()
    if raw == "":
        return None
    if canon_col == "upc":
        return raw  # keep as text — preserves leading zeros
    if canon_col in ("darigold item number", "tp item number", "unit per case"):
        return int(raw) if raw.lstrip("-").isdigit() else raw
    if canon_col in (_PB_OLD, _PB_NEW, _PB_CHANGE):
        try:
            return float(raw)
        except ValueError:
            return None
    if canon_col in ("price start date", "price end date"):
        dt = pd.to_datetime(raw, errors="coerce")
        return dt.to_pydatetime() if pd.notna(dt) else raw
    return raw


def _merge_anchor(ws, row: int, col: int) -> tuple[int, int]:
    """Resolve a cell to its merged-range top-left anchor (writable cell)."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col


def _set_header_value(ws, label: str, value) -> None:
    """Write ``value`` into the cell right of the header-block ``label`` cell."""
    if value is None:
        return
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, min(ws.max_column, 8) + 1):
            cell = ws.cell(r, c)
            if cell.value is not None and _norm_text(cell.value).lower() == label:
                for cc in range(c + 1, ws.max_column + 1):
                    nxt = ws.cell(r, cc)
                    if nxt.value not in (None, ""):
                        ar, ac = _merge_anchor(ws, r, cc)
                        ws.cell(ar, ac).value = value
                        return
    # No existing value cell to the right (rare): nothing to overwrite.


def _clone_and_fill_template(template_bytes: bytes, header_info: dict, rows: list[dict]):
    """Clone the Excel template and populate header + data rows from a PDF.

    Returns ``(wb, ws, header_row, cols)`` ready for the shared update engine.
    Styling of each written cell is copied from the template's first data row so
    the regenerated rows match the template; surplus template rows are deleted.
    """
    from copy import copy
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    _set_header_value(ws, _PB_SITE_LABEL, header_info.get("site"))
    _set_header_value(ws, "customer", header_info.get("customer"))
    _set_header_value(ws, "customer site no", header_info.get("site_no"))

    header = _find_header(ws)
    if header is None:
        raise PriceBookUpdateError("Template has no 'Darigold Item Number' header row.")
    header_row, cols = header
    lo, hi = min(cols.values()), max(cols.values())

    # Capture per-column style from the template's first data row (the look we
    # replicate for every written row).
    ref_style = {}
    for c in range(lo, hi + 1):
        ref = ws.cell(header_row + 1, c)
        ref_style[c] = (copy(ref.font), copy(ref.border), copy(ref.fill),
                        copy(ref.alignment), ref.number_format)

    last_template_row = ws.max_row
    write_r = header_row + 1
    for row in rows:
        for canon, col in cols.items():
            ws.cell(write_r, col, _coerce_pdf_value(canon, row.get(canon, "")))
            font, border, fill, align, numfmt = ref_style[col]
            cell = ws.cell(write_r, col)
            cell.font, cell.border, cell.fill = copy(font), copy(border), copy(fill)
            cell.alignment, cell.number_format = copy(align), numfmt
        write_r += 1

    # Drop any leftover template data rows beyond what we wrote.
    if last_template_row >= write_r:
        ws.delete_rows(write_r, last_template_row - write_r + 1)
    return wb, ws, header_row, cols


def _xlsx_name(pdf_name: str) -> str:
    return re.sub(r"\.pdf$", "", pdf_name, flags=re.IGNORECASE) + ".xlsx"


def update_price_book_from_pdf(
    file_name: str,
    pdf_bytes: bytes,
    template_bytes: bytes,
    lookup: dict,
    known_sites: set,
    old_periods: list,
    new_periods: list,
) -> UpdateReport:
    """Convert a PDF price book to an updated, cloned-template ``.xlsx``."""
    out_name = _xlsx_name(file_name)
    try:
        header_info, rows = _extract_pdf_price_book(pdf_bytes)
    except Exception as exc:  # noqa: BLE001
        return UpdateReport(out_name, False, "error", f"Could not read the PDF: {exc}")
    if not rows:
        return UpdateReport(out_name, False, "error",
                            "No price-book rows found in the PDF (is it a scanned image?).")
    site_raw = header_info.get("site")
    if not site_raw:
        return UpdateReport(out_name, False, "error",
                            "Couldn't read 'Customer Site Name' from the PDF header.")

    try:
        wb, ws, header_row, cols = _clone_and_fill_template(template_bytes, header_info, rows)
    except PriceBookUpdateError as exc:
        return UpdateReport(out_name, False, "error", str(exc),
                            site_name=site_raw, site_found=True)
    except Exception as exc:  # noqa: BLE001
        return UpdateReport(out_name, False, "error",
                            f"Could not build Excel from the template: {exc}",
                            site_name=site_raw, site_found=True)

    missing = [c for c in (_PB_ITEM, _PB_UOM, _PB_OLD, _PB_NEW) if c not in cols]
    if missing:
        return UpdateReport(out_name, False, "error",
                            f"Template is missing column(s): {missing}",
                            site_name=site_raw, site_found=True)

    site = _norm_key(site_raw)
    if site not in known_sites:
        out = io.BytesIO(); wb.save(out)
        return UpdateReport(
            out_name, True, "warning",
            f"Site '{site_raw}' not in the Oracle extract — converted to Excel "
            "but no rows updated.",
            site_name=site_raw, site_found=False, rows_total=len(rows),
            workbook_bytes=out.getvalue(),
        )

    stats = _apply_updates_to_worksheet(
        ws, header_row, cols, site, lookup, old_periods, new_periods)
    out = io.BytesIO(); wb.save(out)
    return _build_update_report(out_name, site_raw, stats, out.getvalue())


def update_price_books(
    files: list[tuple[str, bytes]],
    extract_df: pd.DataFrame,
    old_periods: list,
    new_periods: list,
    *,
    template_bytes: Optional[bytes] = None,
) -> list[UpdateReport]:
    """Update each uploaded price book; route ``.pdf`` → cloned-template Excel.

    ``.xlsx`` inputs are edited in place (formatting preserved). ``.pdf`` inputs
    are converted to an updated copy of the Excel ``template_bytes`` (required —
    a clear error is returned per PDF if it's absent).

    ``old_periods`` / ``new_periods`` are lists of selected snapshots; an empty
    list leaves that price column untouched, so the update targets only the
    period(s) the user selected.
    """
    lookup, sites = build_lookup(extract_df)
    reports: list[UpdateReport] = []
    for name, data in files:
        if name.lower().endswith(".pdf"):
            if template_bytes is None:
                reports.append(UpdateReport(
                    _xlsx_name(name), False, "error",
                    "Upload an Excel template to convert PDF price books."))
            else:
                reports.append(update_price_book_from_pdf(
                    name, data, template_bytes, lookup, sites, old_periods, new_periods))
        else:
            reports.append(update_price_book(name, data, lookup, sites, old_periods, new_periods))
    return reports


def build_zip(reports: list[UpdateReport]) -> bytes:
    """Bundle every workbook that produced bytes into a single ZIP."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for rep in reports:
            if rep.workbook_bytes is not None:
                zf.writestr(f"Updated_{rep.file_name}", rep.workbook_bytes)
    return buf.getvalue()


__all__ = [
    "PriceBookUpdateError",
    "UpdateReport",
    "load_oracle_extract",
    "available_periods",
    "format_period",
    "build_lookup",
    "update_price_book",
    "update_price_book_from_pdf",
    "update_price_books",
    "build_zip",
]
